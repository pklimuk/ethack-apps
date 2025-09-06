import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any
import os
from datetime import datetime
from ..calculators.metrics import PoolMetrics
from ..config.settings import config
from ..utils.logger import logger
from ..utils.helpers import format_currency, format_percentage

class ExcelExporter:
    """Excel report generator for liquidity pool data"""
    
    def __init__(self):
        self.workbook = None
        self.filename = None
        
        # Styling
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def create_comprehensive_report(self, all_metrics: List[PoolMetrics], 
                                  protocol_breakdown: Dict[str, List[PoolMetrics]]) -> str:
        """Create a comprehensive Excel report with multiple sheets"""
        
        # Ensure export directory exists
        os.makedirs(config.EXPORT_PATH, exist_ok=True)
        
        # Generate filename
        self.filename = os.path.join(config.EXPORT_PATH, config.get_export_filename())
        
        # Create workbook
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        logger.info(f"Creating comprehensive Excel report: {self.filename}")
        
        # Create sheets
        self._create_summary_sheet(all_metrics)
        self._create_protocol_sheets(protocol_breakdown)
        self._create_top_performers_sheet(all_metrics)
        self._create_risk_analysis_sheet(all_metrics)
        self._create_comparison_sheet(all_metrics)
        
        if config.INCLUDE_CHARTS:
            self._add_charts_to_summary()
        
        # Save workbook
        self.workbook.save(self.filename)
        logger.info(f"Excel report saved: {self.filename}")
        
        return self.filename
    
    def _create_summary_sheet(self, all_metrics: List[PoolMetrics]):
        """Create main summary sheet"""
        ws = self.workbook.create_sheet("📊 Summary Dashboard", 0)
        
        # Create DataFrame
        df = pd.DataFrame([
            {
                'Protocol': m.protocol,
                'Network': m.network,
                'Pool': m.pool_name,
                'TVL (USD)': m.tvl_usd,
                'Volume 24h (USD)': m.volume_24h,
                'Base APY (%)': m.apy_base,
                'Total APY (%)': m.apy_total,
                'Fees 24h (USD)': m.fees_24h,
                'IL 7d (%)': m.impermanent_loss_7d,
                'Risk Score': m.risk_score,
                'Sharpe Ratio': m.sharpe_ratio or 0,
                'Price Impact 1%': f"{m.price_impact_1pct:.2f}%",
                'Liquidity Depth': m.liquidity_depth,
                'Pool Address': m.pool_address[:10] + '...'
            }
            for m in all_metrics
        ])
        
        # Sort by TVL descending
        df = df.sort_values('TVL (USD)', ascending=False)
        
        # Add header info
        ws['A1'] = "LIQUIDITY POOLS COMPARISON REPORT"
        ws['A1'].font = Font(size=16, bold=True, color='366092')
        ws.merge_cells('A1:N1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:N2')
        
        ws['A3'] = f"Total Pools Analyzed: {len(all_metrics)} | Total TVL: {format_currency(df['TVL (USD)'].sum())}"
        ws['A3'].font = Font(size=12, bold=True)
        ws.merge_cells('A3:N3')
        
        # Write data starting from row 5
        start_row = 5
        for r_idx, (_, row) in enumerate(df.iterrows(), start_row):
            for c_idx, (col, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format cells
                if col in ['TVL (USD)', 'Volume 24h (USD)', 'Fees 24h (USD)']:
                    if isinstance(value, (int, float)) and value > 0:
                        cell.value = format_currency(value)
                elif col in ['Base APY (%)', 'Total APY (%)']:
                    if isinstance(value, (int, float)):
                        cell.value = format_percentage(value)
                elif col in ['IL 7d (%)']:
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.2f}%"
                elif col == 'Risk Score':
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.1f}"
        
        # Add headers
        headers = list(df.columns)
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row-1, column=c_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Apply formatting
        self._apply_conditional_formatting(ws, start_row, len(df), headers)
        self._auto_adjust_columns(ws)
        
        logger.info(f"Created summary sheet with {len(df)} pools")
    
    def _create_protocol_sheets(self, protocol_breakdown: Dict[str, List[PoolMetrics]]):
        """Create individual sheets for each protocol"""
        
        protocol_icons = {
            'Uniswap V3': '🦄',
            'SushiSwap': '🍣', 
            'Curve': '🌊',
            'PancakeSwap': '🥞'
        }
        
        for protocol, pools in protocol_breakdown.items():
            if not pools:
                continue
                
            icon = protocol_icons.get(protocol, '📈')
            sheet_name = f"{icon} {protocol}"
            ws = self.workbook.create_sheet(sheet_name)
            
            # Create detailed DataFrame for this protocol
            df = self._create_detailed_protocol_df(pools)
            
            # Add protocol-specific header
            ws['A1'] = f"{protocol.upper()} LIQUIDITY POOLS"
            ws['A1'].font = Font(size=14, bold=True, color='366092')
            ws.merge_cells('A1:P1')
            
            ws['A2'] = f"Pools: {len(pools)} | Avg APY: {df['Total APY (%)'].mean():.2f}% | Total TVL: {format_currency(df['TVL (USD)'].sum())}"
            ws.merge_cells('A2:P2')
            
            # Write data
            start_row = 4
            self._write_dataframe_to_sheet(ws, df, start_row)
            self._auto_adjust_columns(ws)
            
            logger.info(f"Created {protocol} sheet with {len(pools)} pools")
    
    def _create_detailed_protocol_df(self, pools: List[PoolMetrics]) -> pd.DataFrame:
        """Create detailed DataFrame for protocol-specific sheets"""
        return pd.DataFrame([
            {
                'Rank': i + 1,
                'Pool Name': p.pool_name,
                'Network': p.network,
                'Token0': p.token0_symbol,
                'Token1': p.token1_symbol,
                'Token0 Price': format_currency(p.token0_price) if p.token0_price > 0 else 'N/A',
                'Token1 Price': format_currency(p.token1_price) if p.token1_price > 0 else 'N/A',
                'Token0 Reserve': f"{p.token0_reserve:,.2f}",
                'Token1 Reserve': f"{p.token1_reserve:,.2f}",
                'TVL (USD)': p.tvl_usd,
                'Volume 24h': p.volume_24h,
                'Base APR (%)': p.apr_base,
                'Rewards APR (%)': p.apr_rewards,
                'Total APY (%)': p.apy_total,
                'IL 1d (%)': p.impermanent_loss_1d,
                'IL 7d (%)': p.impermanent_loss_7d,
                'Risk Score': p.risk_score,
                'Sharpe Ratio': p.sharpe_ratio or 0,
                'Pool Address': p.pool_address
            }
            for i, p in enumerate(sorted(pools, key=lambda x: x.tvl_usd, reverse=True))
        ])
    
    def _create_top_performers_sheet(self, all_metrics: List[PoolMetrics]):
        """Create top performers analysis sheet"""
        ws = self.workbook.create_sheet("🏆 Top Performers")
        
        # Top 10 by different metrics
        top_by_apy = sorted(all_metrics, key=lambda x: x.apy_total, reverse=True)[:10]
        top_by_tvl = sorted(all_metrics, key=lambda x: x.tvl_usd, reverse=True)[:10]
        top_by_volume = sorted(all_metrics, key=lambda x: x.volume_24h, reverse=True)[:10]
        best_risk_adjusted = sorted([m for m in all_metrics if m.sharpe_ratio], 
                                  key=lambda x: x.sharpe_ratio, reverse=True)[:10]
        
        # Create sections
        current_row = 1
        
        sections = [
            ("TOP 10 BY APY", top_by_apy, ['Pool', 'Protocol', 'APY (%)', 'TVL (USD)', 'Risk Score']),
            ("TOP 10 BY TVL", top_by_tvl, ['Pool', 'Protocol', 'TVL (USD)', 'APY (%)', 'Volume 24h']),
            ("TOP 10 BY VOLUME", top_by_volume, ['Pool', 'Protocol', 'Volume 24h', 'APY (%)', 'TVL (USD)']),
            ("BEST RISK-ADJUSTED RETURNS", best_risk_adjusted, ['Pool', 'Protocol', 'Sharpe Ratio', 'APY (%)', 'Risk Score'])
        ]
        
        for section_title, pools, columns in sections:
            # Section header
            ws.cell(row=current_row, column=1, value=section_title).font = Font(size=12, bold=True)
            current_row += 1
            
            # Column headers
            for c_idx, col in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
            current_row += 1
            
            # Data rows
            for pool in pools:
                row_data = self._get_section_row_data(pool, columns)
                for c_idx, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=c_idx, value=value)
                current_row += 1
            
            current_row += 2  # Space between sections
        
        self._auto_adjust_columns(ws)
        logger.info("Created top performers sheet")
    
    def _create_risk_analysis_sheet(self, all_metrics: List[PoolMetrics]):
        """Create risk analysis sheet"""
        ws = self.workbook.create_sheet("⚠️ Risk Analysis")
        
        # Risk categories
        low_risk = [m for m in all_metrics if m.risk_score <= 30]
        medium_risk = [m for m in all_metrics if 30 < m.risk_score <= 60]
        high_risk = [m for m in all_metrics if m.risk_score > 60]
        
        # Summary statistics
        ws['A1'] = "RISK ANALYSIS SUMMARY"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        current_row = 3
        
        # Risk distribution
        risk_data = [
            ['Risk Category', 'Pool Count', 'Avg APY (%)', 'Avg TVL (USD)', 'Avg IL 7d (%)'],
            ['Low Risk (0-30)', len(low_risk), 
             sum(p.apy_total for p in low_risk) / len(low_risk) if low_risk else 0,
             sum(p.tvl_usd for p in low_risk) / len(low_risk) if low_risk else 0,
             sum(p.impermanent_loss_7d for p in low_risk) / len(low_risk) if low_risk else 0],
            ['Medium Risk (31-60)', len(medium_risk),
             sum(p.apy_total for p in medium_risk) / len(medium_risk) if medium_risk else 0,
             sum(p.tvl_usd for p in medium_risk) / len(medium_risk) if medium_risk else 0,
             sum(p.impermanent_loss_7d for p in medium_risk) / len(medium_risk) if medium_risk else 0],
            ['High Risk (61+)', len(high_risk),
             sum(p.apy_total for p in high_risk) / len(high_risk) if high_risk else 0,
             sum(p.tvl_usd for p in high_risk) / len(high_risk) if high_risk else 0,
             sum(p.impermanent_loss_7d for p in high_risk) / len(high_risk) if high_risk else 0]
        ]
        
        for row_data in risk_data:
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=value)
                if current_row == 3:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
            current_row += 1
        
        self._auto_adjust_columns(ws)
        logger.info("Created risk analysis sheet")
    
    def _create_comparison_sheet(self, all_metrics: List[PoolMetrics]):
        """Create protocol comparison sheet"""
        ws = self.workbook.create_sheet("📈 Protocol Comparison")
        
        # Group by protocol
        protocol_stats = {}
        for pool in all_metrics:
            protocol = pool.protocol
            if protocol not in protocol_stats:
                protocol_stats[protocol] = []
            protocol_stats[protocol].append(pool)
        
        # Create comparison DataFrame
        comparison_data = []
        for protocol, pools in protocol_stats.items():
            comparison_data.append({
                'Protocol': protocol,
                'Pool Count': len(pools),
                'Total TVL (USD)': sum(p.tvl_usd for p in pools),
                'Avg APY (%)': sum(p.apy_total for p in pools) / len(pools),
                'Avg Risk Score': sum(p.risk_score for p in pools) / len(pools),
                'Total Volume 24h': sum(p.volume_24h for p in pools),
                'Best Pool APY (%)': max(p.apy_total for p in pools),
                'Avg IL 7d (%)': sum(p.impermanent_loss_7d for p in pools) / len(pools)
            })
        
        df = pd.DataFrame(comparison_data)
        df = df.sort_values('Total TVL (USD)', ascending=False)
        
        # Write to sheet
        ws['A1'] = "PROTOCOL COMPARISON"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:I1')
        
        self._write_dataframe_to_sheet(ws, df, 3)
        self._auto_adjust_columns(ws)
        
        logger.info("Created protocol comparison sheet")
    
    def _get_section_row_data(self, pool: PoolMetrics, columns: List[str]) -> List:
        """Get row data for top performers sections"""
        data_map = {
            'Pool': pool.pool_name,
            'Protocol': pool.protocol,
            'APY (%)': f"{pool.apy_total:.2f}%",
            'TVL (USD)': format_currency(pool.tvl_usd),
            'Risk Score': f"{pool.risk_score:.1f}",
            'Volume 24h': format_currency(pool.volume_24h),
            'Sharpe Ratio': f"{pool.sharpe_ratio:.2f}" if pool.sharpe_ratio else "N/A"
        }
        return [data_map.get(col, '') for col in columns]
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, start_row: int):
        """Write DataFrame to worksheet"""
        # Headers
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=c_idx, value=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data
        for r_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    def _apply_conditional_formatting(self, ws, start_row: int, num_rows: int, headers: List[str]):
        """Apply conditional formatting to enhance readability"""
        
        # APY columns - green scale
        apy_cols = [i for i, h in enumerate(headers, 1) if 'APY' in h]
        for col in apy_cols:
            range_str = f"{chr(64+col)}{start_row}:{chr(64+col)}{start_row + num_rows - 1}"
            rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                                mid_type='percentile', mid_value=50, mid_color='92D050',
                                end_type='max', end_color='00B050')
            ws.conditional_formatting.add(range_str, rule)
        
        # Risk Score - red scale (higher = more red)
        risk_col = next((i for i, h in enumerate(headers, 1) if 'Risk Score' in h), None)
        if risk_col:
            range_str = f"{chr(64+risk_col)}{start_row}:{chr(64+risk_col)}{start_row + num_rows - 1}"
            rule = ColorScaleRule(start_type='min', start_color='00B050',
                                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                end_type='max', end_color='FF0000')
            ws.conditional_formatting.add(range_str, rule)
        
        # TVL data bars
        tvl_col = next((i for i, h in enumerate(headers, 1) if 'TVL' in h), None)
        if tvl_col:
            range_str = f"{chr(64+tvl_col)}{start_row}:{chr(64+tvl_col)}{start_row + num_rows - 1}"
            rule = DataBarRule(start_type='min', end_type='max',
                             color='366092', showValue=True, minLength=0, maxLength=90)
            ws.conditional_formatting.add(range_str, rule)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        try:
            for column in ws.columns:
                max_length = 0
                # Skip merged cells by checking if first cell has column_letter
                try:
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value and hasattr(cell, 'column_letter'):
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
                except:
                    # Skip merged cells or cells without column_letter
                    continue
        except Exception as e:
            # If auto-adjustment fails, just continue
            logger.warning(f"Column auto-adjustment failed: {e}")
    
    def _add_charts_to_summary(self):
        """Add charts to summary sheet"""
        try:
            ws = self.workbook["📊 Summary Dashboard"]
            
            # This would add charts - simplified for now due to complexity
            # Could add bar charts for top pools, pie charts for protocol distribution, etc.
            logger.info("Chart creation skipped - would require additional chart logic")
            
        except Exception as e:
            logger.warning(f"Could not add charts: {e}")

# Global exporter instance
excel_exporter = ExcelExporter()